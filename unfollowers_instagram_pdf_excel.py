#!/usr/bin/env python3
# pip install reportlab openpyxl
import json
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
import openpyxl
from openpyxl.styles import Font, PatternFill

# Read the JSON files of followers and followings
with open('followers_1.json', 'r') as file:
    followers_data = json.load(file)

with open('following.json', 'r') as file:
    followings_data = json.load(file)

# Create a vector with their nicknames, accordingly to the structure of the files
followers = [entry['value'] for item in followers_data for entry in item['string_list_data']]
followings = [entry['value'] for item in followings_data['relationships_following'] for entry in item['string_list_data']]

# Create a PDF document
pdf_file = "UnfollowersInstagram.pdf"
doc = SimpleDocTemplate(pdf_file, pagesize=letter)
styles = getSampleStyleSheet()

# Custom styles for PDF
highlight_style = ParagraphStyle(
    name="Highlight",
    parent=styles['Normal'],
    fontName="Helvetica-Bold",
    fontSize=13,
    backColor=colors.orange
)

normal_style = ParagraphStyle(
    name="Normal",
    parent=styles['Normal'],
    fontName="Helvetica",
    fontSize=12
)

# Create an Excel workbook
excel_file = "UnfollowersInstagram.xlsx"
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Unfollowers & Fans"

# Custom styles for Excel
highlight_font = Font(name="Helvetica", bold=True, size=15)
highlight_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
normal_font = Font(name="Helvetica", size=14)

def unfollowers(followings, followers, story, sheet):
    # PDF Part
    story.append(Paragraph("You follow them but they do not follow you back:", highlight_style))
    story.append(Spacer(1, 12))
    
    # Excel Part
    sheet.append(["You follow them but they do not follow you back:"])
    # Fill with highlighter until the end of the text, which is at column 7
    for col in range(1, 8):
        cell = sheet.cell(row=sheet.max_row, column=col)
        cell.font = highlight_font
        cell.fill = highlight_fill
    
    tot = 0
    for i in followings:
        if i not in followers:
            tot += 1
            # PDF Part
            story.append(Paragraph(f"{tot} {i}", normal_style))
            # Excel Part
            sheet.append([f"{tot} {i}"])
            cell = sheet.cell(row=sheet.max_row, column=1)
            cell.font = normal_font
    
    # PDF Part
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Total: {tot}", highlight_style))
    
    # Excel Part
    sheet.append([f"Total: {tot}"])
    # Fill with highlighter until the end of the text, which is at column 2
    cell = sheet.cell(row=sheet.max_row, column=1)
    cell.font = highlight_font
    cell.fill = highlight_fill
    cell = sheet.cell(row=sheet.max_row, column=2)
    cell.font = highlight_font
    cell.fill = highlight_fill
    sheet.append([]) # Empty row for spacing

def fans(followings, followers, story, sheet):
    # PDF Part
    story.append(Paragraph("They follow you but you do not follow them:", highlight_style))
    story.append(Spacer(1, 12))
    
    # Excel Part
    row = sheet.max_row + 1
    sheet.append(["They follow you but you do not follow them:"])
    for col in range(1, 7):
        cell = sheet.cell(row=sheet.max_row, column=col)
        cell.font = highlight_font
        cell.fill = highlight_fill
    
    tot = 0
    for i in followers:
        if i not in followings:
            tot += 1
            # PDF Part
            story.append(Paragraph(f"{tot} {i}", normal_style))
            # Excel Part
            sheet.append([f"{tot} {i}"])
            cell = sheet.cell(row=sheet.max_row, column=1)
            cell.font = normal_font
    
    # PDF Part
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Total: {tot}", highlight_style))
    
    # Excel Part
    sheet.append([f"Total: {tot}"])
    # Fill with highlighter until the end of the text, which is at column 2
    cell = sheet.cell(row=sheet.max_row, column=1)
    cell.font = highlight_font
    cell.fill = highlight_fill
    cell = sheet.cell(row=sheet.max_row, column=2)
    cell.font = highlight_font
    cell.fill = highlight_fill

# Build the PDF content
story = []

unfollowers(followings, followers, story, sheet)
fans(followings, followers, story, sheet)

# Generate the PDF
doc.build(story)
print(f"PDF report generated: {pdf_file}")
# Save the Excel file
workbook.save(excel_file)
print(f"Excel report generated: {excel_file}")